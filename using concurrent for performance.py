import openpyxl
from FileLoader import FileLoader, FilterBy
import logging
from datetime import datetime
from FixedWidthTextParser.Parser import Parser
import os
import concurrent.futures

class Benevity(FileLoader):
    headerParser = None
    transactionParser = None
    trailerParser = None
    transaction = None

    def __init__(self, name, log: logging.Logger, startDate, endDate) -> None:
        super().__init__(name, log, '1900-01-01', endDate)  # Replacing startDate for Benevity

        # Setting up the loader
        self.matching_tables_to_clean = ['BENEVITY_DMS']
        self.stat_queries = {
            'UNMATCHED_STATS': """
                SELECT 'Benevity' AS SOURCE, DonationDate AS TRANSACTION_DATE, MERCHANT_ID,
                SUM(TotalDonationToBeAcknowledged) AS AMOUNT, COUNT(*) AS COUNT
                FROM TRUST.BENEVITY WITH (NOLOCK)
                GROUP BY DonationDate, MERCHANT_ID
                """,
            'UNMATCHED': """
                SELECT 'BENEVITY-DMS' AS SOURCE, TRANSACTION_DATE, MERCHANT_ID,
                '' AS TRANSACTION_TIME, '' AS REQUEST_ID, '' AS MERCHANT_REF_NBR,
                '' AS RECONCILIATION_ID, '' AS DMS_FINANCIAL_ID, NULL AS APG_ID, 
                '' AS APP_NAME
                FROM TRUST.BENEVITY_DMS WITH (NOLOCK)
                """,
            'STATS': """
                SELECT DonationDate AS TRANSACTION_DATE, 'BENEVITY' AS SOURCE,
                COUNT(*) AS CNT, SUM(TOTALDONATIONTOBEACKNOWLEDGED) AS AMOUNT
                FROM TRUST.BENEVITY WITH (NOLOCK)
                GROUP BY DonationDate
                """
        }

        self.file_folder = "Benevity_test"
        self.filter_by = FilterBy.MODIFIED_TIME
        self.can_use_s3 = True

    def trim(self):
        self.log.info("Benevity won't trim")

    def check_file_processed(self, file_name):
        conn = self.db_conn(self.sql_server, self.sql_working_database, self.sql_working_username, self.sql_working_password)
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT LoadDate FROM TRUST.ProcessedFiles WHERE FileName = ?", (file_name,))
            return cursor.fetchone() is not None  # Returns True if the file has been processed
        finally:
            cursor.close()
            conn.close()

    def log_processed_file(self, file_name):
        conn = self.db_conn(self.sql_server, self.sql_working_database, self.sql_working_username, self.sql_working_password)
        cursor = conn.cursor()
        try:
            cursor.execute("INSERT INTO TRUST.ProcessedFiles (FileName, LoadDate) VALUES (?, ?)", 
                           (file_name, datetime.now()))
            conn.commit()
        finally:
            cursor.close()
            conn.close()

    def process_all_files(self):
        files = [f for f in os.listdir(self.file_folder) if f.endswith('.xlsx')]
        conn = self.db_conn(self.sql_server, self.sql_working_database, self.sql_working_username, self.sql_working_password)

        try:
            with concurrent.futures.ThreadPoolExecutor() as executor:
                futures = {executor.submit(self.process_single_file, conn, file): file for file in files}
                for future in concurrent.futures.as_completed(futures):
                    file_name = futures[future]
                    try:
                        future.result()
                    except Exception as e:
                        self.log.error(f"Error processing file {file_name}: {e}")
        finally:
            conn.close()

    def process_single_file(self, conn, file_name):
        # Check if the file has already been processed
        if self.check_file_processed(file_name):
            self.log.info(f"File '{file_name}' has already been processed. Skipping.")
            return 

        cursor = conn.cursor()
        cursor.fast_executemany = True
        recordCount = 0
        tuples = []
        
        file_path = os.path.join(self.file_folder, file_name)
        sql = '''
            INSERT INTO TRUST.BENEVITY (
                COMPANY, PROJECT, DONATIONDATE, FIRSTNAME, LASTNAME, EMAIL, ADDRESS, CITY,
                STATECODE, ZIPCODE, ACTIVITY, COMMENT, TRANSACTIONID, DONATIONFREQUENCY, CURRENCY,
                PROJECTREMOTEID, SOURCE, REASON, TOTALDONATIONTOBEACKNOWLEDGED, MATCHAMOUNT,
                CAUSESUPPORTFEE, MERCHANT_FEE, FEECOMMENT
            ) VALUES (
                ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
            )
        '''
        
        try:
            file = openpyxl.load_workbook(file_path)

            # Find the sheet whose name starts with DonationReport
            donation_report_sheet = next((file[sheet_name] for sheet_name in file.sheetnames if sheet_name.startswith('DonationReport')), None)
            if not donation_report_sheet:
                self.log.info("No sheet found with name starting with 'DonationReport', skipping file")
                return
            
            print("Processing File", file_name)
            for row in donation_report_sheet.iter_rows(min_row=2, max_row=donation_report_sheet.max_row, values_only=True):
                tuple_data = (
                    row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9], 
                    row[10], row[11], row[12], row[13], row[14], row[15], row[16], row[17], row[18], 
                    row[19], row[20], row[21], row[22]
                )
                tuples.append(tuple_data)
                recordCount += 1

                if len(tuples) >= self.sql_batch_size:
                    cursor.executemany(sql, tuples)
                    conn.commit()
                    tuples = []

            if len(tuples) > 0:
                cursor.executemany(sql, tuples)
                conn.commit()

            self.log.info(f"Finished Benevity File '{file_name}' with {recordCount} records.")
            self.log_processed_file(file_name)  # Log the processed file
        except Exception as e:
            conn.rollback()
            self.log.error(f"Benevity Loader: Error inserting records into database: {repr(e)}")
        finally:
            cursor.close()

    def get_matchers(self, matchDate):
        return {
            'Benevity->DMS': {
                'sql': """
                    INSERT INTO TRUST.BENEVITY_DMS (MERCHANT_ID, COMPANY, TRANSACTION_DATE, AMOUNT)
                    SELECT MERCHANT_ID, COMPANY, TRANSACTION_DATE, amount
                    FROM (
                        SELECT MERCHANT_ID, COMPANY, DONATIONDATE AS TRANSACTION_DATE, SUM(TOTALDONATIONTOBEACKNOWLEDGED) AS amount 
                        FROM TRUST.BENEVITY WITH (NOLOCK)
                        GROUP BY COMPANY, DONATIONDATE, MERCHANT_ID
                    ) B
                    WHERE B.TRANSACTION_DATE = ?
                    AND NOT EXISTS (
                        SELECT 1 
                        FROM TRUST.DMS D WITH (NOLOCK) 
                        WHERE D.PAYMENTMETHODCODE = 1 
                        AND D.lastname = B.COMPANY
                        AND B.TRANSACTION_DATE = D.TRANSACTION_DATE
                        AND B.amount = D.AMOUNT
                    )
                """,
                'parameters': [matchDate]
            }
        }

    def filter_out_file_name(self, file_path) -> bool:
        return "Benevity" not in file_path or "Thumbs.db" in file_path
